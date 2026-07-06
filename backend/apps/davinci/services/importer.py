"""
達文西申報 xlsx 解析 pipeline

流程（開發計畫 5）：
    讀檔 → 表頭定位（前 5 列掃描）→ 欄位名建索引 → 逐列清洗（cleaner）
    → 帳號去重（dedup）→ 七指標計算（indicators）→ 匯入報告

已知格式變體：
- 生醫檔：表頭第 1 列；「醫令序號」（11504）vs「序號」（11505）
- 新竹檔：第 1 列是標題「11505達文西案件名單」，表頭在第 2 列；
  病歷號/姓名未遮罩 → 系統遮罩；尾端有空欄
- 期別權威：費用年月欄（西元 yyyymm），不用分頁名（民國年月）推算
"""
from __future__ import annotations

import io
from dataclasses import dataclass, field
from typing import Any

from openpyxl import load_workbook

from ..constants import CAMPUS_ALIASES, HEADER_ALIASES, HEADER_DETECT_KEYS
from . import cleaner
from .dedup import DedupCase, ParsedRow, dedup_rows
from .indicators import AggregatedValue, compute_indicators
from .masking import mask_chart_no, mask_patient_name

HEADER_SCAN_ROWS = 5  # 表頭最多往下找 5 列


@dataclass
class DavinciParseResult:
    cases: list[DedupCase] = field(default_factory=list)
    values: list[AggregatedValue] = field(default_factory=list)
    rows_raw: int = 0
    report: dict = field(default_factory=lambda: {
        "summary": [],           # 每 (campus, period) 摘要 + 七指標
        "cleaned": [],           # 清洗/近似明細
        "conflicts": [],         # yn_conflict / merged_value_mismatch
        "pending": [],           # 無法歸屬院區、期別錯誤等待人工確認
        "header_warnings": [],   # 表頭缺漏/未知
        "masked": 0,             # 系統遮罩筆數
    })


def _norm_header(s: Any) -> str:
    return str(s or "").replace(" ", "").replace("　", "").strip()


def _find_header(rows: list[tuple]) -> tuple[int, dict[str, int]] | None:
    """掃前 N 列找表頭列，回傳 (列 index, 欄位名 → 欄 index)。"""
    for i, row in enumerate(rows[:HEADER_SCAN_ROWS]):
        texts = [_norm_header(c) for c in row]
        if all(any(key in t for t in texts) for key in HEADER_DETECT_KEYS):
            col_map: dict[str, int] = {}
            for field_name, aliases in HEADER_ALIASES.items():
                for alias in aliases:
                    alias_n = _norm_header(alias)
                    hit = next(
                        (j for j, t in enumerate(texts) if t == alias_n),
                        None,
                    )
                    if hit is None:
                        # 退一步：前綴比對（源檔表頭偶有尾註差異）
                        hit = next(
                            (j for j, t in enumerate(texts)
                             if t and alias_n and t.startswith(alias_n)),
                            None,
                        )
                    if hit is not None:
                        col_map[field_name] = hit
                        break
            return i, col_map
    return None


def _resolve_campus(name_raw: Any, code_raw: Any) -> str | None:
    """院區名稱優先，備援院區代碼（去前導 0）。"""
    name = str(name_raw or "").strip()
    if name in CAMPUS_ALIASES:
        return CAMPUS_ALIASES[name]
    code = str(code_raw or "").strip().lstrip("0")
    if code in CAMPUS_ALIASES:
        return CAMPUS_ALIASES[code]
    return None


# 必要表頭：缺任一則整張分頁列入 pending
_REQUIRED_FIELDS = ("campus_name", "period", "account")


def parse_davinci_workbook(content: bytes, file_name: str = "") -> DavinciParseResult:
    """走訪所有分頁 → 清洗 → 去重 → 計算七指標 → 產出報告。不觸資料庫。"""
    result = DavinciParseResult()
    wb = load_workbook(io.BytesIO(content), data_only=True, read_only=True)

    parsed_rows: list[ParsedRow] = []

    for ws in wb.worksheets:
        rows = [tuple(r) for r in ws.iter_rows(values_only=True)]
        if not rows:
            continue
        found = _find_header(rows)
        if found is None:
            result.report["header_warnings"].append(
                f"分頁「{ws.title}」找不到表頭列（前 {HEADER_SCAN_ROWS} 列無「院區名稱」+「帳號」），已略過"
            )
            continue
        header_idx, col = found
        missing = [f for f in _REQUIRED_FIELDS if f not in col]
        if missing:
            result.report["header_warnings"].append(
                f"分頁「{ws.title}」缺少必要欄位 {missing}，已略過"
            )
            continue
        # 非必要欄位缺漏 → 示警但續行
        for f in HEADER_ALIASES:
            if f not in col and f not in ("campus_code",):
                result.report["header_warnings"].append(
                    f"分頁「{ws.title}」缺欄位「{HEADER_ALIASES[f][0]}」"
                )

        def cell(row: tuple, field_name: str) -> Any:
            j = col.get(field_name)
            return row[j] if j is not None and j < len(row) else None

        for ri in range(header_idx + 1, len(rows)):
            row = rows[ri]
            row_no = ri + 1  # Excel 1-based
            account = str(cell(row, "account") or "").strip()
            if account == "":
                continue  # 空列
            result.rows_raw += 1

            campus = _resolve_campus(cell(row, "campus_name"), cell(row, "campus_code"))
            if campus is None:
                result.report["pending"].append({
                    "sheet": ws.title, "row": row_no,
                    "issue": "campus_unresolved",
                    "detail": f"院區無法歸屬：{cell(row, 'campus_name')} / {cell(row, 'campus_code')}",
                })
                continue

            period = cleaner.clean_period(cell(row, "period"))
            if period is None:
                result.report["pending"].append({
                    "sheet": ws.title, "row": row_no,
                    "issue": "period_invalid",
                    "detail": f"費用年月無法解析：{cell(row, 'period')}",
                })
                continue

            flags: list[str] = []

            def _record_clean(fld: str, raw: Any, cleaned: Any, fl: list[str]) -> None:
                for f in fl:
                    result.report["cleaned"].append({
                        "sheet": ws.title, "row": row_no, "campus": campus,
                        "period": period, "field": fld,
                        "raw": None if raw is None else str(raw),
                        "cleaned": cleaned, "flag": f,
                    })

            blood, fl = cleaner.clean_blood_ml(cell(row, "blood_ml"))
            _record_clean("出血量", cell(row, "blood_ml"), blood, fl)
            flags += fl

            op_time, fl = cleaner.clean_op_time(cell(row, "op_time"))
            _record_clean("手術時間", cell(row, "op_time"), op_time, fl)
            flags += fl

            adverse_codes, adverse_text, fl = cleaner.parse_adverse(cell(row, "adverse_content"))
            flags += fl
            severe_codes, _severe_text, fl = cleaner.parse_severe(cell(row, "severe_content"))
            flags += fl

            conv_reason = str(cell(row, "conversion_reason") or "").strip()

            def _clean_yn_field(fld: str, raw: Any, content_has_value: bool = False) -> bool:
                """Y/N 欄清洗 + 記錄：矛盾進 conflicts、未知值進 cleaned（供人工覆核）。"""
                value, fl_ = cleaner.clean_yn(raw, content_has_value=content_has_value)
                for f in fl_:
                    if f == "yn_conflict_content_wins":
                        result.report["conflicts"].append({
                            "sheet": ws.title, "row": row_no, "campus": campus,
                            "period": period, "field": fld, "flag": f,
                        })
                    elif f == "yn_unrecognized_as_n":
                        _record_clean(fld, raw, "N", [f])
                flags.extend(fl_)
                return value

            adverse = _clean_yn_field(
                "不良事件", cell(row, "adverse_flag"),
                content_has_value=bool(adverse_codes or adverse_text),
            )
            severe = _clean_yn_field(
                "嚴重併發症", cell(row, "severe_flag"),
                content_has_value=bool(severe_codes),
            )
            infection = _clean_yn_field("術後感染", cell(row, "infection_flag"))
            conversion = _clean_yn_field(
                "術中轉換", cell(row, "conversion_flag"),
                content_has_value=bool(conv_reason),
            )
            reoperation = _clean_yn_field("再次手術", cell(row, "reoperation_flag"))

            adm, _, fl = cleaner.clean_date(cell(row, "admission_date"))
            flags += fl
            dis, _, fl = cleaner.clean_date(cell(row, "discharge_date"))
            flags += fl
            op_date, op_date_raw, fl = cleaner.clean_date(cell(row, "op_date"))
            flags += fl

            chart_masked, m1 = mask_chart_no(str(cell(row, "chart_no") or ""))
            name_masked, m2 = mask_patient_name(str(cell(row, "patient_name") or ""))
            if m1 or m2:
                flags.append("masked_by_system")
                result.report["masked"] += 1

            parsed_rows.append(ParsedRow(
                row_no=row_no,
                sheet=ws.title,
                campus=campus,
                period=period,
                account=account,
                chart_no_masked=chart_masked,
                patient_masked=name_masked,
                davinci_type=str(cell(row, "davinci_type") or "").strip(),
                dept_code=str(cell(row, "dept_code") or "").strip(),
                dept_name=str(cell(row, "dept_name") or "").strip(),
                surgeon=str(cell(row, "surgeon") or "").strip(),
                order_code=str(cell(row, "order_code") or "").strip(),
                order_name=str(cell(row, "order_name") or "").strip(),
                admission_date=adm,
                discharge_date=dis,
                op_date=op_date,
                op_date_raw=op_date_raw,
                op_time_min=op_time,
                blood_ml=blood,
                conversion=conversion,
                conversion_reason=conv_reason,
                adverse_14d=adverse,
                adverse_codes=adverse_codes,
                adverse_free_text=adverse_text,
                severe_comp_30d=severe,
                severe_comp_codes=severe_codes,
                infection_14d=infection,
                reoperation_14d=reoperation,
                flags=flags,
            ))

    wb.close()

    result.cases = dedup_rows(parsed_rows)
    result.values = compute_indicators(result.cases)

    # merged_value_mismatch 進 conflicts（去重階段才會出現）
    for case in result.cases:
        if "merged_value_mismatch" in case.flags:
            result.report["conflicts"].append({
                "sheet": "-", "row": case.source_rows, "campus": case.campus,
                "period": case.period, "field": "連續值合併",
                "flag": "merged_value_mismatch",
            })

    # summary：每 (campus, period) 一筆
    by_cp: dict[tuple[str, int], list[DedupCase]] = {}
    for c in result.cases:
        by_cp.setdefault((c.campus, c.period), []).append(c)
    for (campus, period), grp in sorted(by_cp.items()):
        indicator_rows = [
            {
                "code": v.indicator_code,
                "numerator": v.numerator,
                "denominator": v.denominator,
                "value": v.value,
                "median": v.median_value,
                "n_excluded": v.n_excluded,
            }
            for v in result.values
            if v.campus == campus and v.period == period
        ]
        result.report["summary"].append({
            "campus": campus,
            "period": period,
            "period_label": cleaner.period_to_roc_label(period),
            "cases_dedup": len(grp),
            "rows_raw": sum(len(c.source_rows) for c in grp),
            "indicators": indicator_rows,
        })

    return result
