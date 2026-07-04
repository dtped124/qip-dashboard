"""
新竹醫院 Excel 解析器

格式特徵：
- 單一 sheet，名稱含「新竹」
- 橫向時間軸：欄位 = 月份（110年01月 ~ 115年N月），每年含 Q1-Q4 欄
- 比率指標佔 2 行（分子/分母）
- 計數指標為 加總+總計 或 單一行（F='-'）

欄位配置：
A=面向, B=序號, C=代碼, D=指標名稱, E=報表名稱, F=計算公式, G=子代碼, H=子名稱
I 以後 = 數據欄（月份/季度）
"""
from __future__ import annotations

import re
from dataclasses import dataclass, field
from typing import Any

import xlrd
from openpyxl import Workbook

from apps.indicators.constants import INDICATOR_META
from .excel_parser import (
    ParseResult,
    ParsedDataPoint,
    ParsedSubcategoryDataPoint,
    ParsedYearlySummary,
    _weighted_year_avg,
)


# 新竹源檔某些 block 的 G 欄子代碼有 typo（例：HA10-01 子代碼寫成 HA10-10-NN）。
# 一律以「父代碼前綴 + G 欄末兩碼」重組 → 確保子代碼跟 element_schema.json 對得起來。
def _normalize_sub_code(parent_code: str, raw_sub_code: str) -> str | None:
    """從 G 欄抽出末兩碼後與父代碼組成標準子代碼；無末兩碼回 None。"""
    m = re.search(r"-(\d{2})$", raw_sub_code.strip())
    if not m:
        return None
    return f"{parent_code}-{m.group(1)}"


@dataclass
class TimeCol:
    col: int       # 0-based column index
    year: int      # 民國年
    month: int     # 1-12 (monthly) or 0 (quarterly)
    quarter: int   # 0 (monthly) or 1-4 (quarterly)


@dataclass
class IndicatorBlock:
    code: str
    name: str
    start_row: int
    formula_type: str  # 'ratio', 'count_total', 'count_single'
    numerator_row: int
    denominator_row: int | None = None
    total_row: int | None = None
    value_row: int | None = None


def _build_time_columns_xlrd(sheet: xlrd.sheet.Sheet) -> list[TimeCol]:
    """Parse header row to build time column mapping (xlrd)."""
    cols: list[TimeCol] = []
    for c in range(sheet.ncols):
        header = str(sheet.cell_value(0, c)).strip()
        if not header:
            continue

        # Format: "110年01月" / "113年1月"
        m = re.match(r"(\d{3})年(\d{1,2})月", header)
        if m:
            cols.append(TimeCol(col=c, year=int(m.group(1)), month=int(m.group(2)), quarter=0))
            continue

        # Format: "111年Q1" / "113Q2"
        m = re.match(r"(\d{3})年?Q(\d)", header)
        if m:
            cols.append(TimeCol(col=c, year=int(m.group(1)), month=0, quarter=int(m.group(2))))
            continue

        # Format: pure "Q1"
        m = re.match(r"^Q(\d)$", header)
        if m and cols:
            last_year = cols[-1].year
            cols.append(TimeCol(col=c, year=last_year, month=0, quarter=int(m.group(1))))

    return cols


def _build_time_columns_openpyxl(sheet) -> list[TimeCol]:
    """Parse header row to build time column mapping (openpyxl)."""
    cols: list[TimeCol] = []
    header_row = list(sheet.iter_rows(min_row=1, max_row=1, values_only=True))[0]
    for c, val in enumerate(header_row):
        if val is None:
            continue
        header = str(val).strip()

        m = re.match(r"(\d{3})年(\d{1,2})月", header)
        if m:
            cols.append(TimeCol(col=c, year=int(m.group(1)), month=int(m.group(2)), quarter=0))
            continue

        m = re.match(r"(\d{3})年?Q(\d)", header)
        if m:
            cols.append(TimeCol(col=c, year=int(m.group(1)), month=0, quarter=int(m.group(2))))
            continue

        m = re.match(r"^Q(\d)$", header)
        if m and cols:
            last_year = cols[-1].year
            cols.append(TimeCol(col=c, year=last_year, month=0, quarter=int(m.group(1))))

    return cols


def _identify_blocks(rows: list[list], ncols: int) -> list[IndicatorBlock]:
    """Scan rows to identify indicator blocks by code in column C."""
    blocks: list[IndicatorBlock] = []

    for i in range(1, len(rows)):
        row = rows[i]
        code = str(row[2] if len(row) > 2 else "").strip()
        if not re.match(r"^HA\d{2}-\d{2}$", code):
            continue

        name = str(row[3] if len(row) > 3 else "").replace("\n", " ").strip()
        formula = str(row[5] if len(row) > 5 else "").strip()

        if formula == "分子":
            # Ratio indicator: locate denominator row.
            #   策略 1：優先找 G 欄末兩碼為 "02" 的列（最可靠 — 直接對應分母 element code）
            #   策略 2：若無，退回找 F 欄 == "分母" 的列（一般用法）
            # 新竹 HA06-01 來源結構特殊：R37 F="分母" 但 G 欄 typo 標成 HA06-01-01
            # 重複腹膜透析數值；真正的血液透析（HA06-01-02）在 R38 F="-"。
            # 策略 1 能正確抓到 R38。
            denom_row = None
            denom_row_by_g = None
            denom_row_by_f = None
            for j in range(i + 1, len(rows)):
                next_code = str(rows[j][2] if len(rows[j]) > 2 else "").strip()
                if re.match(r"^HA\d{2}-\d{2}$", next_code):
                    break
                raw_g = str(rows[j][6] if len(rows[j]) > 6 else "").strip()
                if denom_row_by_g is None and re.search(r"-02$", raw_g):
                    # 同時也要 G 的前綴吻合（避免抓到下一個指標的子代碼）
                    normalized = _normalize_sub_code(code, raw_g)
                    if normalized == f"{code}-02":
                        denom_row_by_g = j
                next_f = str(rows[j][5] if len(rows[j]) > 5 else "").strip()
                if denom_row_by_f is None and next_f == "分母":
                    denom_row_by_f = j
            denom_row = denom_row_by_g if denom_row_by_g is not None else denom_row_by_f
            blocks.append(IndicatorBlock(
                code=code, name=name, start_row=i,
                formula_type="ratio",
                numerator_row=i, denominator_row=denom_row,
            ))
        elif formula == "加總":
            # Count indicator with subtotals
            total_row = None
            for j in range(i + 1, len(rows)):
                next_code = str(rows[j][2] if len(rows[j]) > 2 else "").strip()
                if re.match(r"^HA\d{2}-\d{2}$", next_code):
                    break
                next_f = re.sub(r"[\s\u3000]+", "", str(rows[j][5] if len(rows[j]) > 5 else ""))
                if next_f == "總計":
                    total_row = j
                    break
            blocks.append(IndicatorBlock(
                code=code, name=name, start_row=i,
                formula_type="count_total",
                numerator_row=i, total_row=total_row,
            ))
        elif formula == "-":
            # Single count value
            blocks.append(IndicatorBlock(
                code=code, name=name, start_row=i,
                formula_type="count_single",
                numerator_row=i, value_row=i,
            ))

    return blocks


def _read_numeric(row: list, col_idx: int) -> float | None:
    """Safely read a numeric value from a row."""
    if col_idx >= len(row):
        return None
    raw = row[col_idx]
    if raw is None or raw == "":
        return None
    if isinstance(raw, (int, float)):
        return float(raw)
    s = str(raw).strip()
    if s in ("NR", "NP", "N/A", "-", ""):
        return None
    try:
        return float(s)
    except ValueError:
        return None


def _extract_subcategories(
    block: IndicatorBlock,
    next_block_start: int,
    rows: list[list],
    monthly_cols: list[TimeCol],
    result: ParseResult,
) -> None:
    """
    從 block 區間內掃描子分類列（G 欄含 element code）→ 寫入
    ParsedSubcategoryDataPoint。
      • HA08-01 / HA10-01：4 / 13 個子分類（formula="加總" 區段）
      • HA06-01：R36(num)/R37(den 重複)/R38(血液透析) — 把 R36+R38 都當子分類
    每月空值就寫 None；不阻斷主流程。
    """
    scan_end = next_block_start if next_block_start > 0 else len(rows)
    # 紀錄已寫過的子代碼避免 HA06-01 R36 / R37 重複（兩列 G 都 HA06-01-01）
    seen: set[str] = set()
    for ri in range(block.start_row, min(scan_end, len(rows))):
        row = rows[ri]
        raw_g = str(row[6] if len(row) > 6 else "").strip()
        sub_code = _normalize_sub_code(block.code, raw_g)
        if not sub_code:
            continue
        if sub_code in seen:
            continue
        seen.add(sub_code)
        for tc in monthly_cols:
            val = _read_numeric(row, tc.col)
            result.subcategory_data_points.append(ParsedSubcategoryDataPoint(
                parent_code=block.code,
                subcategory_code=sub_code,
                campus="新竹",
                year=tc.year,
                month=tc.month,
                value=int(val) if val is not None else None,
            ))


def _process_blocks(
    blocks: list[IndicatorBlock],
    rows: list[list],
    time_cols: list[TimeCol],
) -> ParseResult:
    """Process identified blocks and extract data points."""
    result = ParseResult()
    monthly_cols = [tc for tc in time_cols if tc.month > 0]
    quarterly_cols = [tc for tc in time_cols if tc.quarter > 0]
    all_years = sorted(set(tc.year for tc in monthly_cols))

    result.sheets_processed.append("新竹")

    # 每個 block 後面接哪一列（用來界定子分類掃描範圍）
    block_starts = [b.start_row for b in blocks]
    for idx, block in enumerate(blocks):
        meta = INDICATOR_META.get(block.code)
        if not meta:
            result.errors.append(f"找不到指標元資料: {block.code} ({block.name})")
            continue

        is_quarterly = meta.get("is_quarterly", False)

        if block.formula_type == "ratio" and block.denominator_row is not None:
            num_row = rows[block.numerator_row]
            den_row = rows[block.denominator_row]
            target_cols = quarterly_cols if is_quarterly else monthly_cols

            for tc in target_cols:
                numerator = _read_numeric(num_row, tc.col)
                denominator = _read_numeric(den_row, tc.col)

                value: float | None = None
                if numerator is not None and denominator is not None and denominator > 0:
                    raw_ratio = numerator / denominator
                    if meta["unit"] == "percent":
                        value = raw_ratio * 100
                    elif meta["unit"] == "permille":
                        value = raw_ratio * 1000
                    else:
                        value = raw_ratio
                elif numerator is not None and (denominator is None or denominator == 0):
                    if numerator == 0:
                        value = 0

                month = tc.month if not is_quarterly else [0, 1, 4, 7, 10][tc.quarter]

                # 保留小數精度（HA10-09 護病比的 989.9 / 117.1 等不能截斷）
                result.data_points.append(ParsedDataPoint(
                    indicator_code=block.code,
                    campus="新竹",
                    year=tc.year,
                    month=month,
                    value=value,
                    numerator=float(numerator) if numerator is not None else None,
                    denominator=float(denominator) if denominator is not None else None,
                ))

        elif block.formula_type == "count_total" and block.total_row is not None:
            total_row = rows[block.total_row]
            for tc in monthly_cols:
                value = _read_numeric(total_row, tc.col)
                result.data_points.append(ParsedDataPoint(
                    indicator_code=block.code,
                    campus="新竹",
                    year=tc.year,
                    month=tc.month,
                    value=value,
                ))

        elif block.formula_type == "count_single" and block.value_row is not None:
            val_row = rows[block.value_row]
            for tc in monthly_cols:
                value = _read_numeric(val_row, tc.col)
                result.data_points.append(ParsedDataPoint(
                    indicator_code=block.code,
                    campus="新竹",
                    year=tc.year,
                    month=tc.month,
                    value=value,
                ))
        else:
            if block.formula_type == "ratio":
                result.errors.append(f"{block.code}: 找不到分母行")
            else:
                result.errors.append(f"{block.code}: 找不到總計行")
            continue

        # 子分類細項擷取（只跑「加總型」block：HA08-01 / HA10-01）
        # HA06-01 的 -01/-02 已透過修正後的 numerator_row / denominator_row 走正常路徑
        # 範圍：此 block 起，到下一個 block 起（或檔尾）止
        if block.formula_type == "count_total":
            next_block_start = block_starts[idx + 1] if idx + 1 < len(block_starts) else len(rows)
            _extract_subcategories(block, next_block_start, rows, monthly_cols, result)

        # Build yearly summaries
        for year in all_years:
            year_dps = [
                dp for dp in result.data_points
                if dp.indicator_code == block.code
                and dp.year == year
            ]
            avg = _weighted_year_avg(year_dps)
            result.yearly_summaries.append(ParsedYearlySummary(
                indicator_code=block.code,
                campus="新竹",
                year=year,
                average=avg,
            ))

    return result


# ── Entry points ──

def parse_hsinchu_xlrd(book: xlrd.Book) -> ParseResult:
    """Parse Hsinchu format from xlrd workbook."""
    sheet_name = next((n for n in book.sheet_names() if "新竹" in n), None)
    if not sheet_name:
        return ParseResult(errors=["找不到新竹醫院的工作表"])

    sheet = book.sheet_by_name(sheet_name)
    if sheet.nrows < 2:
        return ParseResult(errors=["工作表無資料列"])

    time_cols = _build_time_columns_xlrd(sheet)
    if not time_cols:
        return ParseResult(errors=["無法解析時間欄位（表頭）"])

    # Convert to list of lists
    rows: list[list] = []
    for r in range(sheet.nrows):
        row = [sheet.cell_value(r, c) for c in range(sheet.ncols)]
        rows.append(row)

    blocks = _identify_blocks(rows, sheet.ncols)
    return _process_blocks(blocks, rows, time_cols)


def parse_hsinchu_openpyxl(wb: Workbook) -> ParseResult:
    """Parse Hsinchu format from openpyxl workbook."""
    sheet_name = next((n for n in wb.sheetnames if "新竹" in n), None)
    if not sheet_name:
        return ParseResult(errors=["找不到新竹醫院的工作表"])

    sheet = wb[sheet_name]
    rows: list[list] = []
    for row in sheet.iter_rows(values_only=True):
        rows.append([v if v is not None else "" for v in row])

    if len(rows) < 2:
        return ParseResult(errors=["工作表無資料列"])

    time_cols = _build_time_columns_openpyxl(sheet)
    if not time_cols:
        return ParseResult(errors=["無法解析時間欄位（表頭）"])

    blocks = _identify_blocks(rows, len(rows[0]) if rows else 0)
    return _process_blocks(blocks, rows, time_cols)
