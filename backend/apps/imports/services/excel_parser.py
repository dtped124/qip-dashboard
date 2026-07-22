"""
QIP Excel 解析引擎

支援兩種格式：
1. 竹北/竹東格式：每個工作表 = 一個年度×一個院區，縱向列出指標
   - 110年格式：類別(0) NO(1) 名稱(2) 月份(3-14) 標竿(15-16)
   - 111-115年格式：類別(0) NO(1) 代碼(2) 名稱(3) 月份(4-15) 年均+標竿(16+)
2. 新竹格式：單一工作表，橫向時間軸 → 委派 hsinchu_parser

從 Next.js excel-parser.ts 完整翻譯，保留所有邏輯路徑。
"""
from __future__ import annotations

import io
import math
import re
from dataclasses import dataclass, field
from typing import Any

import xlrd
from openpyxl import load_workbook

from apps.indicators.constants import INDICATOR_META, NAME_TO_CODE, SUBCATEGORY_DEFS
from apps.indicators.strict_schema import (
    SERIES_STITCH_RULES,
    STRICT_INDICATOR_SCHEMA,
    strict_resolve_code,
    validate_nd_names,
    get_kind,
)
from .data_cleaner import clean_value, clean_value_raw, normalize_monthly_value, normalize_benchmark
from .matching import match_indicator_name


@dataclass
class ParsedDataPoint:
    indicator_code: str
    campus: str
    year: int
    month: int
    value: float | None
    # 分子/分母改為 float — 部分指標 (如 HA10-09 護病比) 原本就是小數
    numerator: float | None = None
    denominator: float | None = None


@dataclass
class ParsedYearlySummary:
    indicator_code: str
    campus: str
    year: int
    average: float | None = None
    benchmark_regional: float | None = None
    benchmark_district: float | None = None


@dataclass
class ParsedSubcategoryDataPoint:
    """子分類細項計數（HA08-01 / HA10-01 等指標的子項）。"""
    parent_code: str
    subcategory_code: str
    campus: str
    year: int
    month: int
    value: int | None


@dataclass
class ParseResult:
    data_points: list[ParsedDataPoint] = field(default_factory=list)
    yearly_summaries: list[ParsedYearlySummary] = field(default_factory=list)
    subcategory_data_points: list[ParsedSubcategoryDataPoint] = field(default_factory=list)
    errors: list[str] = field(default_factory=list)
    sheets_processed: list[str] = field(default_factory=list)


# ── Sheet name parsing ──

def _parse_sheet_name(name: str) -> tuple[int, str | None]:
    """Extract ROC year and campus from sheet name. Returns (year, campus)."""
    year_match = re.search(r"(\d{3})年", name)
    if not year_match:
        # 部分來源檔漏打「年」（如「111竹東」）— 接受開頭的裸三位數年份
        year_match = re.match(r"^(\d{3})(?!\d)", name)
    year = int(year_match.group(1)) if year_match else 0
    if year and not (105 <= year <= 130):
        year = 0

    if "竹北" in name and "竹東" in name:
        return year, None  # Combined sheet, skip
    if "竹北" in name:
        return year, "竹北"
    if "竹東" in name:
        return year, "竹東"
    return year, None


# ── Code resolution ──

CATEGORY_MAPPING = {
    "整體照護": "整體照護", "加護照護": "加護照護", "手術照護": "手術照護",
    "產科照護": "產科照護", "急診照護": "急診照護", "重點照護": "重點照護",
    "感染管制": "感染管制", "用藥安全": "用藥安全", "呼吸照護": "呼吸照護",
    "經營管理": "經營管理",
}


def _resolve_code(raw_code: str, raw_name: str) -> str:
    """
    解析指標代碼 — 嚴格比對版。
    只接受名稱完全相符（strip + 壓縮空白後）的指標。任何「相似」、「包含」、
    「模糊」一律不採用，避免 HA06-24 被誤認為 HA06-21 之類的問題。
    """
    code = strict_resolve_code(raw_code or "", raw_name or "")
    return code or ""


# ── Unified cell reader for xlrd/openpyxl ──

class _XlrdSheetAdapter:
    """Adapter for xlrd sheet to provide uniform interface."""

    def __init__(self, book: xlrd.Book, sheet: xlrd.sheet.Sheet):
        self._book = book
        self._sheet = sheet

    @property
    def nrows(self) -> int:
        return self._sheet.nrows

    @property
    def ncols(self) -> int:
        return self._sheet.ncols

    def cell_value(self, row: int, col: int) -> Any:
        if col >= self._sheet.ncols or row >= self._sheet.nrows:
            return ""
        return self._sheet.cell_value(row, col)

    def cell_for_clean(self, row: int, col: int) -> Any:
        """Get cell value suitable for clean_value_raw, handling % format."""
        if col >= self._sheet.ncols or row >= self._sheet.nrows:
            return ""
        cell = self._sheet.cell(row, col)
        if cell.ctype == xlrd.XL_CELL_NUMBER:
            # Check if formatted as % or ‰
            try:
                xf_index = self._sheet.cell_xf_index(row, col)
                fmt = self._book.format_map.get(xf_index)
                if fmt and fmt.format_str:
                    fmt_str = fmt.format_str
                    if "%" in fmt_str or "‰" in fmt_str:
                        # Return the formatted text instead of raw number
                        # xlrd raw % value is 0.0327 for 3.27%
                        val = cell.value
                        if "%" in fmt_str:
                            return f"{val * 100}%"
                        elif "‰" in fmt_str:
                            return f"{val * 1000}‰"
            except Exception:
                pass
        return cell.value

    def header_value(self, col: int) -> str:
        if col >= self._sheet.ncols:
            return ""
        return str(self._sheet.cell_value(0, col)).strip()

    def header_row(self) -> list[str]:
        return [self.header_value(c) for c in range(self._sheet.ncols)]


class _OpenpyxlSheetAdapter:
    """Adapter for openpyxl sheet."""

    def __init__(self, sheet):
        self._sheet = sheet
        self._rows: list[list] | None = None

    def _ensure_rows(self):
        if self._rows is None:
            self._rows = []
            for row in self._sheet.iter_rows(values_only=False):
                self._rows.append(list(row))

    @property
    def nrows(self) -> int:
        self._ensure_rows()
        return len(self._rows)

    @property
    def ncols(self) -> int:
        return self._sheet.max_column or 0

    def cell_value(self, row: int, col: int) -> Any:
        self._ensure_rows()
        if row >= len(self._rows) or col >= len(self._rows[row]):
            return ""
        cell = self._rows[row][col]
        return cell.value if cell.value is not None else ""

    def cell_for_clean(self, row: int, col: int) -> Any:
        """Get cell value handling % format."""
        self._ensure_rows()
        if row >= len(self._rows) or col >= len(self._rows[row]):
            return ""
        cell = self._rows[row][col]
        if cell.value is None:
            return ""
        if isinstance(cell.value, (int, float)) and cell.number_format:
            fmt = str(cell.number_format)
            if "%" in fmt:
                return f"{cell.value * 100}%"
            if "‰" in fmt:
                return f"{cell.value * 1000}‰"
        return cell.value

    def header_value(self, col: int) -> str:
        return str(self.cell_value(0, col)).strip()

    def header_row(self) -> list[str]:
        return [self.header_value(c) for c in range(self.ncols)]


# ── Column-layout detection ──

# Matches "1月" or "115.1月" / "115-1月" / "115/1月" / "115年1月" but NOT "11月"/"12月"
_FIRST_MONTH_RE = re.compile(r"^(?:\d{2,3}[.\-/年]\s*)?1月$")


def _detect_columns(header_row: list[str], year: int) -> tuple[int, int, int]:
    """Detect (code_col, name_col, month_start) from header row.

    Header keywords:
      - 代碼 → code column (e.g. 'HA01-01')
      - 指標名稱 / 指標定義 / 分子 → name column
      - '1月' / 'YYY.1月' → first month column

    Falls back to year-based defaults if header is ambiguous.
    """
    code_col = -1
    name_col = -1
    month_start = -1

    for c, h in enumerate(header_row):
        hs = str(h).replace("\n", "").replace(" ", "").strip()
        if not hs:
            continue
        if code_col < 0 and hs == "代碼":
            code_col = c
        if name_col < 0 and ("指標名稱" in hs or "指標定義" in hs or "分子" in hs):
            name_col = c
        if month_start < 0 and _FIRST_MONTH_RE.match(hs):
            month_start = c

    # Fallback to year-based defaults if detection fails
    if name_col < 0 or month_start < 0:
        if year == 110:
            return -1, 2, 3
        return 2, 3, 4

    return code_col, name_col, month_start


# ── Main parser ──

def parse_qip_excel(file_bytes: bytes, file_name: str) -> ParseResult:
    """
    主入口：解析 QIP 指標 Excel 檔案。

    支援 .xls (xlrd) 和 .xlsx (openpyxl) 格式。
    自動偵測新竹格式並委派專用解析器。
    """
    result = ParseResult()
    ext = file_name.lower().rsplit(".", 1)[-1] if "." in file_name else ""

    # Open workbook
    sheets: list[tuple[str, Any]] = []  # (sheet_name, adapter)

    if ext == "xls":
        book = xlrd.open_workbook(file_contents=file_bytes, formatting_info=True)
        sheet_names = book.sheet_names()

        # Check for Hsinchu format
        if any("新竹" in name for name in sheet_names):
            from .hsinchu_parser import parse_hsinchu_xlrd
            return parse_hsinchu_xlrd(book)

        for sn in sheet_names:
            sheets.append((sn, _XlrdSheetAdapter(book, book.sheet_by_name(sn))))

    elif ext == "xlsx":
        wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
        sheet_names = wb.sheetnames

        if any("新竹" in name for name in sheet_names):
            from .hsinchu_parser import parse_hsinchu_openpyxl
            return parse_hsinchu_openpyxl(wb)

        for sn in sheet_names:
            sheets.append((sn, _OpenpyxlSheetAdapter(wb[sn])))
    else:
        result.errors.append(f"不支援的檔案格式: .{ext}")
        return result

    nd_computed_keys: set[str] = set()

    # Process each sheet
    for sheet_name, adapter in sheets:
        year, campus = _parse_sheet_name(sheet_name)

        if not campus:
            continue
        if year == 0:
            result.errors.append(f"無法解析工作表年度: {sheet_name}")
            continue

        result.sheets_processed.append(sheet_name)

        if adapter.nrows < 2:
            continue

        # Determine column layout — detect from header (with year-based fallback)
        header_row = adapter.header_row()
        code_col, name_col, month_start = _detect_columns(header_row, year)
        is_110_layout = code_col < 0  # no 代碼 column → 110-style layout
        month_end = month_start + 12

        current_category = ""
        # 本工作表已成功產出資料的指標代碼。用於：以「無項次」放寬條件納入的
        # 變體重複列（如同一 code 的不含PAC 版本），若該 code 已解析過就跳過，
        # 確保放寬項次判斷不會製造重複資料。有項次的列不受此限（維持原行為）。
        emitted_codes: set[str] = set()

        for i in range(1, adapter.nrows):
            # Update category
            cat_raw = str(adapter.cell_value(i, 0)).strip()
            if cat_raw:
                resolved_cat = CATEGORY_MAPPING.get(cat_raw)
                if resolved_cat:
                    current_category = resolved_cat

            # Check if indicator row.
            # 主要判準：Col B (項次) 為正整數。
            # 例外：部分來源檔會漏填項次（如 115 竹東把 HA01-03 拆成 含/不含PAC
            # 後刪掉了項次號），此時只要「代碼」欄有值即視為指標列。
            # 分子/分母等子列的代碼欄一律為空，故不受影響；至於重複的變體列
            # （如不含PAC）雖同樣以代碼解析成同一 code，仍會被後續的分子/分母
            # 嚴格名稱驗證擋下，不會產生重複資料。
            no_val = adapter.cell_value(i, 1)
            has_seq = (
                isinstance(no_val, (int, float))
                and not isinstance(no_val, bool)
                and no_val == int(no_val)
                and no_val > 0
            )
            code_cell = "" if code_col < 0 else str(adapter.cell_value(i, code_col)).strip()
            if not has_seq and not code_cell:
                continue

            # Extract code and name
            raw_code = code_cell
            raw_name = str(adapter.cell_value(i, name_col)).strip()
            code = _resolve_code(raw_code, raw_name)

            if not code:
                seq_label = int(no_val) if has_seq else "-"
                result.errors.append(
                    f"無法解析指標代碼（嚴格比對未通過）: year={year} campus={campus} NO={seq_label} name={raw_name}"
                )
                continue

            # 放寬項次判斷帶來的變體重複列防護：若此列靠「代碼欄有值」而非項次
            # 被納入，且該 code 在本工作表已成功解析過，視為重複變體，跳過。
            if not has_seq and code in emitted_codes:
                continue

            # ── 嚴格驗證：分子/分母列名必須完全相符（rate 類指標）──
            # 對 rate 類指標：先掃描接下來的「空 NO 列」找出分子/分母列名，
            # 跟 STRICT_INDICATOR_SCHEMA 對照。任一不符 → 跳過該指標、寫警告。
            # 註：subcategory / single_value 不做 n/d 驗證（沒有 n/d 結構）。
            _kind = get_kind(code)
            if _kind == "rate":
                # 先掃 分子:/分母: 標記列；找不到就 fallback 用 i+1 / i+2
                def _is_blank_no_row_local(row_idx: int) -> bool:
                    if row_idx >= adapter.nrows:
                        return False
                    v = adapter.cell_value(row_idx, 1)
                    if v == "" or v is None:
                        return True
                    if isinstance(v, str) and v.strip() == "":
                        return True
                    return False

                num_row_for_validation = -1
                den_row_for_validation = -1
                j = i + 1
                scan_end = min(i + 8, adapter.nrows)
                while j < scan_end and _is_blank_no_row_local(j):
                    text = str(adapter.cell_value(j, name_col)).strip()
                    if num_row_for_validation < 0 and re.match(r"^分子\s*[:：]", text):
                        num_row_for_validation = j
                    if den_row_for_validation < 0 and re.match(r"^分母\s*[:：]", text):
                        den_row_for_validation = j
                    j += 1
                # Fallback：沒有顯式標記就用 i+1 / i+2
                if num_row_for_validation < 0 and _is_blank_no_row_local(i + 1):
                    num_row_for_validation = i + 1
                if den_row_for_validation < 0 and _is_blank_no_row_local(i + 2):
                    den_row_for_validation = i + 2

                num_name = str(adapter.cell_value(num_row_for_validation, name_col)).strip() if num_row_for_validation > 0 else None
                den_name = str(adapter.cell_value(den_row_for_validation, name_col)).strip() if den_row_for_validation > 0 else None

                ok, reason = validate_nd_names(code, num_name, den_name)
                if not ok:
                    result.errors.append(
                        f"分子/分母列名未通過嚴格比對 → 跳過: code={code} year={year} campus={campus} {reason}"
                    )
                    continue

            # Step A: Extract n/d from adjacent rows
            # Two layouts supported:
            #   (a) Inline fraction in one row: "10/100" (legacy)
            #   (b) Separated rows: row+1 = numerator values, row+2 = denominator values
            # Both layouts have empty 項次 in the n/d rows.
            adjacent_nd: list[dict | None] = [None] * 12

            def _is_blank_no_row(row_idx: int) -> bool:
                if row_idx >= adapter.nrows:
                    return False
                v = adapter.cell_value(row_idx, 1)
                if v == "" or v is None:
                    return True
                if isinstance(v, str) and v.strip() == "":
                    return True
                return False

            def _coerce_number(v: Any) -> float | None:
                """保留浮點精度（HA10-09 護病比的 549.28、65.77 等不能截斷）。"""
                if isinstance(v, bool):
                    return None
                if isinstance(v, (int, float)) and not isinstance(v, bool):
                    if isinstance(v, float) and math.isnan(v):
                        return None
                    return float(v)
                if isinstance(v, str):
                    s = v.strip()
                    if not s:
                        return None
                    try:
                        return float(s)
                    except ValueError:
                        return None
                return None

            # 別名：呼叫端用同樣名稱（避免改太多行）
            _coerce_int = _coerce_number

            if i + 1 < adapter.nrows and _is_blank_no_row(i + 1):
                # Try (a) inline-fraction layout first
                inline_found = False
                for m in range(12):
                    col_idx = month_start + m
                    nd_str = str(adapter.cell_value(i + 1, col_idx)).strip()
                    frac_match = re.search(r"\(?(\d+)\s*/\s*(\d+)\)?", nd_str)
                    if frac_match:
                        adjacent_nd[m] = {
                            "numerator": int(frac_match.group(1)),
                            "denominator": int(frac_match.group(2)),
                        }
                        inline_found = True

                if not inline_found:
                    # (b1) Try explicit 分子:/分母: markers within the next blank-NO
                    # rows. 竹東 HA05-01 places 分子/分母 on non-adjacent rows
                    # with intermediate breakdown rows ─ scan up to 7 rows ahead.
                    num_row_idx = -1
                    den_row_idx = -1
                    j = i + 1
                    scan_end = min(i + 8, adapter.nrows)
                    while j < scan_end and _is_blank_no_row(j):
                        text = str(adapter.cell_value(j, name_col)).strip()
                        # Must be 分子/分母 followed by a colon (ASCII or full-width)
                        # to avoid matching unrelated labels like "分母案件中…"
                        if num_row_idx < 0 and re.match(r"^分子\s*[:：]", text):
                            num_row_idx = j
                        if den_row_idx < 0 and re.match(r"^分母\s*[:：]", text):
                            den_row_idx = j
                        j += 1

                    if num_row_idx > 0 and den_row_idx > 0:
                        # Explicit markers: pick those specific rows
                        for m in range(12):
                            col_idx = month_start + m
                            n_num = _coerce_int(adapter.cell_value(num_row_idx, col_idx))
                            d_num = _coerce_int(adapter.cell_value(den_row_idx, col_idx))
                            if n_num is not None and d_num is not None:
                                adjacent_nd[m] = {
                                    "numerator": n_num,
                                    "denominator": d_num,
                                }
                    elif _is_blank_no_row(i + 2):
                        # (b2) Default: row+1 = numerator, row+2 = denominator
                        for m in range(12):
                            col_idx = month_start + m
                            n_num = _coerce_int(adapter.cell_value(i + 1, col_idx))
                            d_num = _coerce_int(adapter.cell_value(i + 2, col_idx))
                            if n_num is not None and d_num is not None:
                                adjacent_nd[m] = {
                                    "numerator": n_num,
                                    "denominator": d_num,
                                }

            # Step B: Check if rate indicator
            meta = INDICATOR_META.get(code)
            is_rate = meta is not None and meta.get("data_nature") in ("binomial_rate", "poisson_rate")

            # Step C: Extract monthly values
            for m in range(12):
                col_idx = month_start + m
                # Only use cell_for_clean (% format handling) for rate indicators
                # Count/continuous indicators: read raw value to avoid ×100 on counts
                raw = adapter.cell_for_clean(i, col_idx) if is_rate else adapter.cell_value(i, col_idx)

                cr = clean_value_raw(raw)

                # Merge n/d sources
                numerator = cr.numerator
                denominator = cr.denominator
                if adjacent_nd[m]:
                    numerator = adjacent_nd[m]["numerator"]
                    denominator = adjacent_nd[m]["denominator"]

                # Compute value
                value: float | None = None
                computed_from_nd = False

                if is_rate and numerator is not None and denominator is not None and denominator > 0:
                    raw_ratio = numerator / denominator
                    if meta["unit"] == "percent":
                        value = raw_ratio * 100
                    elif meta["unit"] == "permille":
                        value = raw_ratio * 1000
                    else:
                        value = raw_ratio
                    computed_from_nd = True

                    # Sanity check: 比對主儲存格顯示值與 n/d 計算值，差異 >3 倍時
                    # 寫一筆警告供人工檢視。實務上兩側都可能出錯：
                    #   - n/d 偏大 → 分母漏一位數 typo（如 d=23 應為 526）
                    #   - 主儲存格偏大 → 儲存格單位/格式問題
                    # 沒有可靠的通用規則能自動判別哪邊才對，且自動翻轉會傷到
                    # 其他正規化邏輯產出的合理 n/d 值，所以這裡僅警告、不覆蓋。
                    # 確認後的單筆錯誤請以資料庫更新處理。
                    main_cell_value = normalize_monthly_value(
                        cr.value, code, year, campus, cr.had_symbol
                    )
                    if (
                        main_cell_value is not None
                        and main_cell_value > 0
                        and value is not None
                        and value > 0
                    ):
                        ratio = max(value, main_cell_value) / min(value, main_cell_value)
                        if ratio > 3:
                            direction = "n/d 偏大 (疑似分母 typo)" if value > main_cell_value else "儲存格偏大 (疑似格式錯誤)"
                            result.errors.append(
                                f"[警告] {code} {campus} {year}年{m + 1}月: "
                                f"n/d 計算值 {value:.4f} 與儲存格顯示值 {main_cell_value:.4f} "
                                f"差異 {ratio:.1f} 倍 (n={numerator}, d={denominator}) — {direction}；"
                                f"請人工確認後以資料庫更新修正。"
                            )
                elif is_rate and numerator is not None and numerator == 0:
                    value = 0
                    computed_from_nd = True
                else:
                    value = normalize_monthly_value(cr.value, code, year, campus, cr.had_symbol)

                if computed_from_nd:
                    nd_computed_keys.add(f"{code}_{campus}_{year}_{m + 1}")

                result.data_points.append(ParsedDataPoint(
                    indicator_code=code,
                    campus=campus,
                    year=year,
                    month=m + 1,
                    value=value,
                    numerator=numerator,
                    denominator=denominator,
                ))

            emitted_codes.add(code)

            # Extract year average and benchmarks
            benchmark_regional: float | None = None
            benchmark_district: float | None = None

            if is_110_layout and year == 110 and len(header_row) > 16:
                # Legacy 110 layout: cols 15/16 are benchmark
                benchmark_regional = clean_value(adapter.cell_value(i, 15))
                benchmark_district = clean_value(adapter.cell_value(i, 16))
            else:
                after_months = month_end
                # Search headers for benchmark columns
                for c in range(after_months + 1, len(header_row)):
                    h = header_row[c].replace("\n", "")
                    if "區域醫院" in h and benchmark_regional is None:
                        benchmark_regional = clean_value(adapter.cell_value(i, c))
                    elif "地區醫院" in h and benchmark_district is None:
                        benchmark_district = clean_value(adapter.cell_value(i, c))

                # Fallback: try fixed positions
                if benchmark_regional is None and len(header_row) > after_months + 3:
                    last_col = len(header_row) - 1
                    second_last = last_col - 1
                    h_last = header_row[last_col].replace("\n", "")
                    h_second = header_row[second_last].replace("\n", "")
                    if "區域" in h_second or "標竿" in h_second:
                        benchmark_regional = clean_value(adapter.cell_value(i, second_last))
                    if "地區" in h_last or "標竿" in h_last:
                        benchmark_district = clean_value(adapter.cell_value(i, last_col))

            # Normalize benchmarks
            benchmark_regional = normalize_benchmark(benchmark_regional, code, year, campus)
            benchmark_district = normalize_benchmark(benchmark_district, code, year, campus)

            result.yearly_summaries.append(ParsedYearlySummary(
                indicator_code=code,
                campus=campus,
                year=year,
                average=None,  # Recalculate after outlier validation
                benchmark_regional=benchmark_regional,
                benchmark_district=benchmark_district,
            ))

            # ── 子分類細項擷取（HA08-01 / HA10-01）──
            # 主指標列下方接著 N 個子分類列（NO 欄空、依固定順序）。
            # 每列 12 個月份欄各自獨立計數，寫入 ParsedSubcategoryDataPoint。
            # 來源 Excel 子分類數不夠時就讀到沒空白 NO 列為止，靜默忽略。
            sub_codes = SUBCATEGORY_DEFS.get(code)
            if sub_codes:
                for idx, sub_code in enumerate(sub_codes):
                    sub_row = i + 1 + idx
                    if sub_row >= adapter.nrows:
                        break
                    # 必須是空 NO 列才認為是子分類列
                    sub_no = adapter.cell_value(sub_row, 1)
                    if sub_no not in (None, "") and not (isinstance(sub_no, str) and sub_no.strip() == ""):
                        break
                    for m in range(12):
                        col_idx = month_start + m
                        raw = adapter.cell_value(sub_row, col_idx)
                        sub_val: int | None = None
                        if isinstance(raw, (int, float)) and not isinstance(raw, bool):
                            if not (isinstance(raw, float) and math.isnan(raw)):
                                sub_val = int(raw)
                        elif isinstance(raw, str) and raw.strip():
                            try:
                                sub_val = int(float(raw.strip()))
                            except ValueError:
                                sub_val = None
                        result.subcategory_data_points.append(ParsedSubcategoryDataPoint(
                            parent_code=code,
                            subcategory_code=sub_code,
                            campus=campus,
                            year=year,
                            month=m + 1,
                            value=sub_val,
                        ))

    # Outlier validation
    _validate_outliers(result, nd_computed_keys)

    # Series stitching (e.g. HA01-03-01 ≤113 年沿用 HA01-03 竹東歷史)
    _apply_series_stitch(result)

    # Recalculate year averages from corrected monthly data
    _recalculate_year_averages(result)

    return result


def _apply_series_stitch(result: ParseResult) -> None:
    """依 SERIES_STITCH_RULES 將來源指標的歷史資料點複製給新指標。

    執行時機夾在 _validate_outliers 之後（複製的是校正後的值，與來源
    數列完全相同）、_recalculate_year_averages 之前（縫接年度的年均
    自動由複製點重算）。標竿一律 None：不含PAC 無官方標竿，且來源
    標竿屬含PAC 定義，不適用。
    """
    for target_code, rule in SERIES_STITCH_RULES.items():
        src_code = rule["from_code"]
        campus = rule["campus"]
        up_to = rule["up_to_year"]

        existing_dp = {
            (dp.year, dp.month)
            for dp in result.data_points
            if dp.indicator_code == target_code and dp.campus == campus
        }
        cloned_years: set[int] = set()
        for dp in list(result.data_points):
            if (
                dp.indicator_code == src_code
                and dp.campus == campus
                and dp.year <= up_to
                and (dp.year, dp.month) not in existing_dp
            ):
                result.data_points.append(ParsedDataPoint(
                    indicator_code=target_code,
                    campus=campus,
                    year=dp.year,
                    month=dp.month,
                    value=dp.value,
                    numerator=dp.numerator,
                    denominator=dp.denominator,
                ))
                cloned_years.add(dp.year)

        existing_ys = {
            s.year for s in result.yearly_summaries
            if s.indicator_code == target_code and s.campus == campus
        }
        for y in sorted(cloned_years - existing_ys):
            result.yearly_summaries.append(ParsedYearlySummary(
                indicator_code=target_code,
                campus=campus,
                year=y,
                average=None,  # 由 _recalculate_year_averages 重算
                benchmark_regional=None,
                benchmark_district=None,
            ))


def _validate_outliers(result: ParseResult, nd_computed_keys: set[str]) -> None:
    """Detect and auto-correct outliers (translated from validateOutliers in TS)."""
    OUTLIER_THRESHOLD = 20
    CORRECTION_TOLERANCE = 5

    # Group data points by (code, campus)
    groups: dict[str, list[ParsedDataPoint]] = {}
    for dp in result.data_points:
        key = f"{dp.indicator_code}_{dp.campus}"
        groups.setdefault(key, []).append(dp)

    for key, dps in groups.items():
        code = dps[0].indicator_code
        meta = INDICATOR_META.get(code)
        if not meta or meta.get("unit") not in ("percent", "permille"):
            continue

        values = [dp.value for dp in dps if dp.value is not None and dp.value > 0]
        if len(values) < 3:
            continue

        sorted_vals = sorted(values)
        median = sorted_vals[len(sorted_vals) // 2]
        if median == 0:
            continue

        for dp in dps:
            if dp.value is None or dp.value == 0:
                continue

            nd_key = f"{dp.indicator_code}_{dp.campus}_{dp.year}_{dp.month}"
            was_nd = nd_key in nd_computed_keys
            ratio = dp.value / median

            if ratio > OUTLIER_THRESHOLD:
                if was_nd:
                    result.errors.append(
                        f"⚠ 異常值（n/d 計算）: {code} {dp.campus} {dp.year}年{dp.month}月 "
                        f"值={dp.value} (中位數={median:.4f}，為中位數的{ratio:.1f}倍)"
                    )
                else:
                    corrected = dp.value / 100
                    corrected_ratio = corrected / median
                    if 1 / CORRECTION_TOLERANCE <= corrected_ratio <= CORRECTION_TOLERANCE:
                        original = dp.value
                        dp.value = corrected
                        result.errors.append(
                            f"✅ 自動修正: {code} {dp.campus} {dp.year}年{dp.month}月 "
                            f"{original} → {corrected} (÷100)"
                        )
                    else:
                        result.errors.append(
                            f"⚠ 異常值: {code} {dp.campus} {dp.year}年{dp.month}月 "
                            f"值={dp.value} (中位數={median:.4f}，為中位數的{ratio:.1f}倍)"
                        )
            elif ratio < 1 / OUTLIER_THRESHOLD:
                src = "（n/d 計算）" if was_nd else ""
                result.errors.append(
                    f"⚠ 異常值{src}: {code} {dp.campus} {dp.year}年{dp.month}月 "
                    f"值={dp.value} (僅為中位數的{ratio * 100:.2f}%)"
                )


def _weighted_year_avg(dps: list) -> float | None:
    """分母加權平均：sum(value×d)/sum(d)，等價於 sum(n)/sum(d)×scale。
    若無分母資訊則退回算術平均。"""
    valid = [dp for dp in dps if dp.value is not None]
    if not valid:
        return None
    with_den = [dp for dp in valid if dp.denominator and dp.denominator > 0]
    if with_den:
        den_sum = sum(dp.denominator for dp in with_den)
        return sum(dp.value * dp.denominator for dp in with_den) / den_sum
    return sum(dp.value for dp in valid) / len(valid)


def _recalculate_year_averages(result: ParseResult) -> None:
    """Recalculate year averages from monthly data (after outlier correction)."""
    for summary in result.yearly_summaries:
        year_dps = [
            dp for dp in result.data_points
            if dp.indicator_code == summary.indicator_code
            and dp.campus == summary.campus
            and dp.year == summary.year
        ]
        avg = _weighted_year_avg(year_dps)
        if avg is not None:
            summary.average = avg
